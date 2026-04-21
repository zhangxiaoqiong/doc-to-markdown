"""XLSX 数据完整性检测脚本。

检测 openpyxl 兼容性、合并单元格、标题缺失、跨行数据等问题。
返回问题数（0 = 通过），供转换脚本做前置校验。
"""

import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path


def detect_issues(file_path: str) -> int:
    """检测 xlsx 文件的数据问题。返回问题数（0 = 通过）。"""
    issue_count = 0
    print(f"=== 数据完整性检测: {Path(file_path).name} ===\n")

    # 1. 检测 openpyxl 兼容性
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        print(f"[OK] openpyxl 加载成功")
        print(f"  Sheet 数量: {len(wb.sheetnames)}")
        print(f"  Sheet 名称: {', '.join(wb.sheetnames)}")
        openpyxl_ok = True
        wb.close()
    except Exception as e:
        print(f"[FAIL] openpyxl 加载失败: {e}")
        print(f"  → 文件不兼容，无法转换")
        openpyxl_ok = False

    print()

    # 2. 解析 shared strings + 逐 sheet 分析
    strings = []
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            # Shared strings
            with z.open('xl/sharedStrings.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                for si in root:
                    text_parts = []
                    for t in si:
                        if t.text:
                            text_parts.append(t.text)
                        if t.tail:
                            text_parts.append(t.tail)
                    strings.append(''.join(text_parts))

            # Workbook info
            workbook_xml = z.read('xl/workbook.xml')
            wb_tree = ET.fromstring(workbook_xml)
            sheets = wb_tree.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet')
            sheet_count = len(sheets)

            print(f"[OK] XML 解析成功")
            print(f"  Shared strings: {len(strings)} 条")
            print(f"  Sheet 数量: {sheet_count}")
            for s in sheets:
                print(f"  Sheet: {s.get('name')}")
            print()

            # 3. 逐 sheet 分析
            ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
            for sheet_idx, sheet_elem in enumerate(sheets, 1):
                sheet_name = sheet_elem.get('name', f'Sheet{sheet_idx}')
                try:
                    with z.open(f'xl/worksheets/sheet{sheet_idx}.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        rows = root.findall(f'.//{ns}row')

                        if not rows:
                            print(f"  [{sheet_name}] 无数据")
                            continue

                        print(f"  [{sheet_name}] 共 {len(rows)} 行 (含表头)")

                        # 两遍扫描：先找到最后一个非空行的位置，跳过尾部连续空行
                        data_rows = rows[1:]
                        last_non_empty_idx = -1
                        parsed_rows = []
                        for idx, row in enumerate(data_rows):
                            cells = row.findall(f'.//{ns}c')
                            row_data = []
                            for cell in cells:
                                v = cell.find(f'.//{ns}v')
                                if v is not None and v.text:
                                    if cell.get('t') == 's':
                                        try:
                                            idx_s = int(v.text)
                                            row_data.append(strings[idx_s] if idx_s < len(strings) else '')
                                        except:
                                            row_data.append(v.text)
                                    else:
                                        row_data.append(v.text)
                                else:
                                    row_data.append('')
                            parsed_rows.append((row, row_data))
                            col0 = row_data[0] if len(row_data) > 0 else ''
                            col1 = row_data[1] if len(row_data) > 1 else ''
                            col2 = row_data[2] if len(row_data) > 2 else ''
                            if col0 or col1 or col2:
                                last_non_empty_idx = idx

                        # 第二遍：只检查到最后一个非空行
                        total_rows = 0
                        valid_rows = 0
                        empty_rows = 0
                        missing_q = 0
                        missing_a = []
                        potential_merged = []

                        for idx, (row, row_data) in enumerate(parsed_rows):
                            if idx > last_non_empty_idx:
                                break  # 跳过尾部连续空行

                            total_rows += 1
                            col0 = row_data[0] if len(row_data) > 0 else ''
                            col1 = row_data[1] if len(row_data) > 1 else ''
                            col2 = row_data[2] if len(row_data) > 2 else ''

                            if not col0 and not col1 and not col2:
                                empty_rows += 1
                                issue_count += 1
                                continue

                            if not col1:
                                if col2:
                                    potential_merged.append((total_rows, col2[:50]))
                                    issue_count += 1
                                else:
                                    missing_q += 1
                                    issue_count += 1
                                continue

                            # 检查答案
                            has_answer = bool(col2) if len(row_data) >= 3 else True
                            if not has_answer:
                                missing_a.append(col1[:50])
                            else:
                                valid_rows += 1

                        print(f"    数据行(不含尾部空行): {total_rows} 行")
                        print(f"    有效条目: {valid_rows}")
                        if empty_rows:
                            print(f"    [FAIL] 数据中间的空行: {empty_rows} 行")
                        if missing_q:
                            print(f"    [FAIL] 缺少标题: {missing_q} 行")
                        if missing_a:
                            print(f"    缺少答案 ({len(missing_a)} 行，不阻塞转换):")
                            for q in missing_a:
                                print(f"      - {q}")
                        if potential_merged:
                            print(f"    [FAIL] 疑似合并单元格: {len(potential_merged)} 行")
                        print()
                except Exception as e:
                    print(f"  [FAIL] 解析 {sheet_name} 失败: {e}")
                    issue_count += 1
                    print()

    except Exception as e:
        print(f"[FAIL] XML 解析失败: {e}")
        return 1

    # 4. 如果 openpyxl 可用，检测合并单元格
    if openpyxl_ok:
        wb = openpyxl.load_workbook(file_path)
        print("合并单元格检测:")
        for name in wb.sheetnames:
            ws = wb[name]
            merged = list(ws.merged_cells.ranges)
            if merged:
                print(f"  [FAIL] {name}: {len(merged)} 个合并单元格")
                for mc in merged[:5]:
                    print(f"    - {mc}")
                if len(merged) > 5:
                    print(f"    ... 还有 {len(merged) - 5} 个")
                issue_count += len(merged)
            else:
                print(f"  [{name}] 无合并单元格")
        print()
        wb.close()

    print(f"=== 检测完成 === (共发现 {issue_count} 个问题)")
    return issue_count


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python detect_issues.py <input.xlsx>")
        sys.exit(1)
    count = detect_issues(sys.argv[1])
    sys.exit(1 if count > 0 else 0)
