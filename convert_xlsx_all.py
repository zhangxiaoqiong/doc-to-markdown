#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel to Markdown converter - optimized for Dify knowledge base segmentation
Each record is self-contained with complete category information
Removes HTML tags and optimizes for semantic search
"""

import sys
import os
import warnings
import re
from pathlib import Path
import openpyxl
from typing import Tuple, Optional, List

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


def clean_html_tags(text: str) -> str:
    """Remove HTML tags and entities from text."""
    if not text:
        return ""

    # Replace common HTML entities
    text = text.replace('&nbsp;', ' ')
    text = text.replace('&lt;', '<')
    text = text.replace('&gt;', '>')
    text = text.replace('&amp;', '&')
    text = text.replace('<br/>', '\n')
    text = text.replace('<br>', '\n')
    text = text.replace('</br>', '')

    # Remove HTML tags
    text = re.sub(r'<[^>]+>', '', text)

    # Clean multiple spaces
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()

    return text


def merge_multiline_values(values_list: List[str]) -> str:
    """Merge multiple line values into one, useful for keywords."""
    if not values_list:
        return ""

    # Clean each value
    cleaned = [str(v).strip() for v in values_list if v and str(v).strip()]

    if len(cleaned) == 1:
        return cleaned[0]

    # Join with Chinese punctuation or comma
    return '、'.join(cleaned)


def excel_to_markdown(file_path: str) -> str:
    """Convert Excel to Dify-optimized markdown.

    Each record contains:
    - Title (## ...)
    - Category (from first column)
    - All field values from header
    - Separated by ---

    This ensures each block is self-contained and can be properly segmented.
    """

    try:
        import pandas as pd
        df = pd.read_excel(file_path, header=None)

        if df.empty or len(df) < 2:
            return ""

        # Extract header from first row
        header_row = df.iloc[0]
        headers = []
        for h in header_row:
            if pd.notna(h) and h != 'nan':
                headers.append(str(h).strip())
            else:
                headers.append("")

        # Remove trailing empty headers
        while headers and not headers[-1]:
            headers.pop()

        if not headers:
            return ""

        markdown_lines = []

        # Process data rows starting from row 2 (index 1)
        for idx in range(1, len(df)):
            row = df.iloc[idx]

            # Check if entire row is empty
            if all(pd.isna(row[i]) or row[i] == 'nan' for i in range(len(headers))):
                continue

            # Extract values from each column
            values = []
            for col_idx in range(len(headers)):
                try:
                    if col_idx < len(row):
                        val = row[col_idx]
                        if pd.notna(val) and val != 'nan':
                            val_str = str(val).strip()
                            values.append(val_str)
                        else:
                            values.append("")
                    else:
                        values.append("")
                except:
                    values.append("")

            # Skip if no meaningful data
            if not any(values):
                continue

            # ===== Build record =====
            category = values[0] if len(values) > 0 else ""
            title = values[1] if len(values) > 1 else ""

            # Title is required
            if not title:
                continue

            # Add title as ## heading
            markdown_lines.append(f"## {title}")
            markdown_lines.append("")

            # Add category if exists (important for Dify segmentation)
            if category and category != 'None':
                header_0 = headers[0] if headers[0] else "分类"
                markdown_lines.append(f"**{header_0}:** {category}")

            # Add other fields (from column 3 onwards, as columns 1-2 are category/title)
            # But also include columns that represent different data
            for col_idx in range(2, len(headers)):
                if col_idx < len(values) and values[col_idx]:
                    val_str = values[col_idx].strip()

                    # Skip if value is same as title or category
                    if val_str and val_str != title and val_str != category and val_str != 'None':
                        field_name = headers[col_idx] if col_idx < len(headers) else f"字段{col_idx+1}"

                        # Clean HTML tags from field values
                        val_str = clean_html_tags(val_str)

                        if val_str:
                            markdown_lines.append(f"**{field_name}:** {val_str}")

            # Add separator
            markdown_lines.append("")
            markdown_lines.append("---")
            markdown_lines.append("")

        if markdown_lines:
            return '\n'.join(markdown_lines)

        return ""

    except Exception as e:
        pass

    # Fallback: use openpyxl
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True, rich_text=False)
    except:
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except:
            return ""

    markdown_lines = []

    try:
        if not workbook.sheetnames:
            return ""

        for sheet_name in workbook.sheetnames[:1]:
            try:
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
            except:
                continue

            if len(rows) < 2:
                continue

            # Extract header
            header_row = rows[0]
            headers = []
            for h in header_row:
                h_str = str(h).strip() if h and str(h) != 'None' else ""
                headers.append(h_str)

            while headers and not headers[-1]:
                headers.pop()

            if not headers:
                continue

            # Process data rows
            data_rows = rows[1:]

            for row in data_rows:
                if not row or all(cell is None for cell in row):
                    continue

                # Extract values
                values = []
                for col_idx in range(len(headers)):
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val and str(val) != 'None':
                            values.append(str(val).strip())
                        else:
                            values.append("")
                    else:
                        values.append("")

                if not any(values):
                    continue

                # Build record
                category = values[0] if len(values) > 0 else ""
                title = values[1] if len(values) > 1 else ""

                if not title:
                    continue

                # Add title
                markdown_lines.append(f"## {title}")
                markdown_lines.append("")

                # Add category
                if category and category != 'None':
                    header_0 = headers[0] if headers[0] else "分类"
                    markdown_lines.append(f"**{header_0}:** {category}")

                # Add other fields
                for col_idx in range(2, len(headers)):
                    if col_idx < len(values) and values[col_idx]:
                        val_str = values[col_idx].strip()
                        if val_str and val_str != title and val_str != category and val_str != 'None':
                            field_name = headers[col_idx] if col_idx < len(headers) else f"字段{col_idx+1}"
                            val_str = clean_html_tags(val_str)
                            if val_str:
                                markdown_lines.append(f"**{field_name}:** {val_str}")

                # Add separator
                markdown_lines.append("")
                markdown_lines.append("---")
                markdown_lines.append("")

    finally:
        try:
            workbook.close()
        except:
            pass

    return '\n'.join(markdown_lines) if markdown_lines else ""


def convert_xlsx_single_file(input_dir: Path, output_dir: Path, file_path: Path) -> Tuple[bool, Optional[str]]:
    """Convert a single XLSX file to markdown."""
    try:
        if not file_path.exists():
            return False, f"文件不存在: {file_path}"

        markdown_content = excel_to_markdown(str(file_path))

        if not markdown_content.strip():
            return False, "转换结果为空"

        output_file = output_dir / (file_path.stem + '.md')
        output_file.write_text(markdown_content, encoding='utf-8')

        return True, None

    except Exception as e:
        return False, f"转换失败: {str(e)}"


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python convert_xlsx_all.py <input_dir> <output_dir> [file_path]")
        sys.exit(1)

    input_dir = Path(sys.argv[1])
    output_dir = Path(sys.argv[2])

    # If file_path provided, use it directly (it may be relative path or absolute)
    if len(sys.argv) >= 4:
        file_arg = sys.argv[3]
        # If it's a relative path within input_dir, construct full path
        file_path = Path(file_arg)
        if not file_path.is_absolute():
            # Check if it's relative to input_dir
            full_path = input_dir / file_arg
            if full_path.exists():
                file_path = full_path
    else:
        print("Error: file_path is required")
        sys.exit(1)

    success, error = convert_xlsx_single_file(input_dir, output_dir, file_path)

    if success:
        print(f"✓ Successfully converted: {file_arg}")
        sys.exit(0)
    else:
        print(f"✗ Error: {error}")
        sys.exit(1)
