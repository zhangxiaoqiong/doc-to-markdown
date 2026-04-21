"""Simple Excel to Markdown converter - core logic only."""

from pathlib import Path
import openpyxl
import zipfile
import xml.etree.ElementTree as ET
import os
import sys


def generate_row_description(headers: list, row_data: list) -> str:
    """Generate natural language description for a single table row using Claude API."""
    try:
        from anthropic import Anthropic

        auth_token = os.getenv("ANTHROPIC_AUTH_TOKEN")
        base_url = os.getenv("ANTHROPIC_BASE_URL")

        if not auth_token:
            return ""

        client = Anthropic(api_key=auth_token, base_url=base_url)

        field_descriptions = []
        for header, value in zip(headers, row_data):
            header_str = str(header).strip() if header else ""
            value_str = str(value).strip() if value else ""
            if header_str and value_str:
                field_descriptions.append(f"{header_str}：{value_str}")

        if not field_descriptions:
            return ""

        prompt = "根据以下信息，用一句话陈述事实，不需要总结：\n" + "，".join(field_descriptions)

        message = client.messages.create(
            model="claude/claude-haiku-4-5",
            max_tokens=8000,
            messages=[{"role": "user", "content": prompt}]
        )

        return message.content[0].text.strip()
    except Exception as e:
        print(f"警告: API 生成描述失败 ({headers[0] if headers else 'unknown'}): {e}", file=sys.stderr)
        return ""


def _parse_xlsx_from_xml(file_path: str) -> list:
    """Fallback: Parse XLSX using XML extraction for corrupted files.

    Returns list of rows with data.
    """
    all_sheets_data = _parse_all_xlsx_sheets_from_xml(file_path)
    # Return first non-empty sheet for backward compatibility
    for sheet_data in all_sheets_data:
        if sheet_data:
            return sheet_data
    return []


def _parse_all_xlsx_sheets_from_xml(file_path: str) -> list:
    """Parse all sheets from XLSX using XML extraction.

    Returns list of sheet data, where each sheet is a list of rows.
    """
    all_sheets = []

    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            # Get shared strings (same for all sheets)
            strings = []
            try:
                with z.open('xl/sharedStrings.xml') as f:
                    tree = ET.parse(f)
                    root = tree.getroot()
                    for si in root:
                        text_parts = []
                        for t in si:
                            if t.text:
                                text_parts.append(t.text)
                        strings.append(''.join(text_parts))
            except Exception as e:
                print(f"警告: XML fallback 解析 sharedStrings 失败: {e}", file=sys.stderr)

            # Get workbook to find sheet list
            try:
                workbook_xml = z.read('xl/workbook.xml')
                wb_tree = ET.fromstring(workbook_xml)
                sheets = wb_tree.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet')
                sheet_count = len(sheets)
            except Exception as e:
                print(f"警告: XML fallback 无法读取 workbook.xml: {e}", file=sys.stderr)
                sheet_count = 10  # Try up to 10 sheets

            # Parse each sheet
            for sheet_idx in range(1, sheet_count + 1):
                try:
                    with z.open(f'xl/worksheets/sheet{sheet_idx}.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        rows = root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row')
                        sheet_data = []

                        for row in rows:
                            cells = row.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c')
                            row_data = []
                            for cell in cells:
                                v_elem = cell.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                                if v_elem is not None and v_elem.text:
                                    # Check if it's a string reference
                                    if cell.get('t') == 's':
                                        try:
                                            idx = int(v_elem.text)
                                            row_data.append(strings[idx] if idx < len(strings) else '')
                                        except:
                                            row_data.append(v_elem.text)
                                    else:
                                        row_data.append(v_elem.text)
                                else:
                                    row_data.append('')
                            if any(row_data):  # Only add non-empty rows
                                sheet_data.append(row_data)

                        if sheet_data:
                            all_sheets.append(sheet_data)
                except Exception as e:
                    print(f"警告: XML fallback 解析 sheet{sheet_idx} 失败: {e}", file=sys.stderr)

    except Exception as e:
        print(f"警告: XML fallback 解析 ZIP 文件失败: {e}", file=sys.stderr)
        return []

    return all_sheets


def excel_to_markdown_by_sheets(file_path: str, enable_row_descriptions: bool = False) -> dict:
    """Convert Excel to markdown, returning separate markdown for each sheet.

    Returns dict: {sheet_name: markdown_content}
    """
    sheet_markdowns = {}

    # Try openpyxl first
    try:
        workbook = openpyxl.load_workbook(file_path)
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                if rows and len(rows) > 1:  # Only process sheets with data
                    markdown_lines = _process_rows(rows, sheet_name, enable_row_descriptions)
                    if markdown_lines:
                        sheet_markdowns[sheet_name] = '\n'.join(markdown_lines)
        finally:
            workbook.close()
    except Exception as e:
        print(f"警告: openpyxl 加载失败，将使用 XML fallback（可能丢失合并单元格数据）: {e}", file=sys.stderr)
        # Fallback to XML parsing - process all sheets
        all_rows_list = _parse_all_xlsx_sheets_from_xml(file_path)

        # Get sheet names from workbook
        try:
            with zipfile.ZipFile(file_path, 'r') as z:
                workbook_xml = z.read('xl/workbook.xml')
                wb_tree = ET.fromstring(workbook_xml)
                sheets = wb_tree.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet')
                sheet_names = [sheet.get('name', f'Sheet{i+1}') for i, sheet in enumerate(sheets)]
        except Exception as e2:
            print(f"警告: XML fallback 无法获取 sheet 名称，使用默认命名: {e2}", file=sys.stderr)
            sheet_names = [f'Sheet{i+1}' for i in range(len(all_rows_list))]

        for sheet_idx, rows in enumerate(all_rows_list):
            if rows and len(rows) > 1:
                sheet_name = sheet_names[sheet_idx] if sheet_idx < len(sheet_names) else f'Sheet{sheet_idx + 1}'
                markdown_lines = _process_rows(rows, sheet_name, enable_row_descriptions)
                if markdown_lines:
                    sheet_markdowns[sheet_name] = '\n'.join(markdown_lines)

    return sheet_markdowns


def excel_to_markdown(file_path: str) -> str:
    """Convert Excel to simple markdown.

    Reads all rows from all sheets and formats them hierarchically.
    Fallback to XML parsing if openpyxl fails.
    """
    all_markdown_lines = []

    # Try openpyxl first
    try:
        workbook = openpyxl.load_workbook(file_path)
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                if rows and len(rows) > 1:  # Only process sheets with data
                    markdown_lines = _process_rows(rows, sheet_name, False)
                    if markdown_lines:
                        all_markdown_lines.extend(markdown_lines)
        finally:
            workbook.close()
    except Exception as e:
        print(f"警告: openpyxl 加载失败，将使用 XML fallback（可能丢失合并单元格数据）: {e}", file=sys.stderr)
        # Fallback to XML parsing - process all sheets
        all_rows_list = _parse_all_xlsx_sheets_from_xml(file_path)
        for sheet_idx, rows in enumerate(all_rows_list):
            if rows and len(rows) > 1:
                markdown_lines = _process_rows(rows, f"Sheet{sheet_idx + 1}", False)
                if markdown_lines:
                    all_markdown_lines.extend(markdown_lines)

    return '\n'.join(all_markdown_lines) if all_markdown_lines else ""


def _process_rows(rows, sheet_name, enable_row_descriptions=False, max_rows_with_descriptions=50):
    """Process a list of rows and return markdown lines.

    If enable_row_descriptions is True:
    - Output table header + row + description for each data row
    - Description is generated by Claude API (only for first max_rows_with_descriptions rows to control cost)
    - Remaining rows output table format only, no descriptions

    Otherwise:
    - Use original formatting (Q&A or standard table)
    """
    markdown_lines = []

    if not rows or len(rows) < 2:
        return markdown_lines

    # Get header row
    header_row = rows[0]
    data_rows = rows[1:]

    # If descriptions enabled, use new table format for all sheets
    if enable_row_descriptions:
        clean_headers = [str(h).strip().replace('\n', '<br>').replace('\r', '') if h else "" for h in header_row]

        for row in data_rows:
            if not row or all(cell is None for cell in row):
                continue

            clean_row_data = [str(cell).strip().replace('\n', '<br>').replace('\r', '') if cell else "" for cell in row]

            if not any(clean_row_data):
                continue

            # Output: field-value pairs (definition list format)
            markdown_lines.append("")
            for header, value in zip(clean_headers, clean_row_data):
                if header and value:
                    markdown_lines.append(f"**{header}:** {value}")

            # Generate description via API (逐行调用)
            description = generate_row_description(clean_headers, clean_row_data)
            if description:
                markdown_lines.append("")
                markdown_lines.append(description)

            markdown_lines.append("")
            markdown_lines.append("---")

        return markdown_lines

    # Original logic: descriptions disabled
    # Determine if this is Q&A format (3 columns: category, question, answer)
    # or table format (multiple columns with headers)
    is_qa_format = len(header_row) >= 3 and all(header_row[i] for i in range(3))

    # Original logic: Q&A format or descriptions disabled
    current_category = None

    # Statistics for skipped rows
    empty_rows = 0
    missing_col1 = 0
    missing_answer = []

    for row in data_rows:
        if not row or all(cell is None for cell in row):
            empty_rows += 1
            continue

        # Extract first 3 columns for common processing
        col0 = str(row[0]).strip().replace('\n', '<br>').replace('\r', '') if row[0] else ""
        col1 = str(row[1]).strip().replace('\n', '<br>').replace('\r', '') if len(row) > 1 and row[1] else ""
        col2 = str(row[2]).strip().replace('\n', '<br>').replace('\r', '') if len(row) > 2 and row[2] else ""

        if not col1:  # Need at least column 1 (title/question)
            missing_col1 += 1
            continue

        # Q&A format: requires col2 (answer)
        if is_qa_format and not col2:
            missing_answer.append(col1)
            continue

        # Add category heading if changed
        if col0 and col0 != current_category:
            markdown_lines.append(f"# {col0}")
            current_category = col0

        # Add title/question as heading
        markdown_lines.append(f"## {col1}")
        markdown_lines.append("")

        # Add all fields with their header labels
        for col_idx, header in enumerate(header_row):
            if col_idx < len(row) and row[col_idx]:
                cell_val = str(row[col_idx]).strip().replace('\n', '<br>').replace('\r', '')
                if cell_val and cell_val != 'None':
                    header_str = str(header).strip() if header else f"字段{col_idx+1}"
                    markdown_lines.append(f"**{header_str}:** {cell_val}")

        markdown_lines.append("")
        markdown_lines.append("---")
        markdown_lines.append("")

    # Print filtering statistics
    if empty_rows:
        print(f"  [{sheet_name}] 跳过空行: {empty_rows}", file=sys.stderr)
    if missing_col1:
        print(f"  [{sheet_name}] 跳过缺少标题的行(可能为合并单元格跨行): {missing_col1}", file=sys.stderr)
    if missing_answer:
        print(f"  [{sheet_name}] 跳过缺少答案的 {len(missing_answer)} 行:", file=sys.stderr)
        for q in missing_answer:
            print(f"    - {q}", file=sys.stderr)

    return markdown_lines


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        print("Usage: python convert_xlsx.py <input.xlsx> <output_dir|output.md> [--descriptions]")
        print("  --descriptions: 生成带一句话描述的行，输出为目录（多 sheet 时分文件）")
        sys.exit(1)

    input_file = sys.argv[1]
    output_target = sys.argv[2]
    enable_descriptions = "--descriptions" in sys.argv

    # Step 1: 数据完整性验证（有问题直接终止）
    print(f"\n=== 步骤1: 数据完整性验证 ===\n")
    from detect_issues import detect_issues

    issues_found = detect_issues(input_file)
    if issues_found:
        print(f"\n[终止] 数据验证失败，共发现 {issues_found} 个问题，请修正后重试。", file=sys.stderr)
        sys.exit(1)

    print(f"\n=== 步骤2: 开始转换 ===\n")

    # Step 2: 转换
    if enable_descriptions:
        result = excel_to_markdown_by_sheets(input_file, enable_row_descriptions=True)
        output_dir = Path(output_target)
        output_dir.mkdir(parents=True, exist_ok=True)
        base_name = Path(input_file).stem
        for sheet_name, md_content in result.items():
            output_filename = f"{base_name}_{sheet_name}_带描述.md"
            output_file = output_dir / output_filename
            output_file.write_text(md_content, encoding='utf-8')
            print(f"[OK] Generated: {output_file}")
        if not result:
            print("错误: 转换结果为空，请检查上方的警告信息", file=sys.stderr)
            sys.exit(1)
    else:
        output_file = Path(output_target)
        md = excel_to_markdown(input_file)
        if not md:
            print("错误: 转换结果为空，请检查上方的警告信息", file=sys.stderr)
            sys.exit(1)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        output_file.write_text(md, encoding='utf-8')
        print(f"[OK] Generated: {output_file}")
